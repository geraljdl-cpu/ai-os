#!/usr/bin/env python3
"""
export_excel.py — Gera ficheiros .xlsx para exportação de dados operacionais.

Usage:
  python3 bin/export_excel.py <type> [output_path]
  Types: insurance | ideas | decisions

Output: escreve .xlsx para output_path (ou stdout em bytes se omitido via API)
"""
import sys, os, io, json
from datetime import datetime, date

_bin_dir = os.path.dirname(os.path.abspath(__file__))
if _bin_dir in sys.path:
    sys.path.remove(_bin_dir)

_env_file = "/etc/aios.env"
if os.path.exists(_env_file):
    for _l in open(_env_file):
        _l = _l.strip()
        if _l and not _l.startswith("#") and "=" in _l:
            _k, _, _v = _l.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())

import sqlalchemy as sa
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://aios_user:jdl@127.0.0.1:5432/aios")

# ── Estilo ────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
HEADER_FONT  = Font(bold=True, color="F1F5F9", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATE_FORMAT  = "DD/MM/YYYY"


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[row].height = 22


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _fmt(v):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return v


def _conn():
    return sa.create_engine(DATABASE_URL, pool_pre_ping=True, future=True)


# ── Sheets ────────────────────────────────────────────────────────────────────

def sheet_insurance(wb):
    ws = wb.active
    ws.title = "Apólices"
    headers = ["ID", "Entidade", "Ref", "Seguradora", "Nº Apólice", "Categoria",
               "Cobertura", "Início", "Fim", "Renovação", "Frequência",
               "Prémio (€)", "Estado", "Auto-renovar", "Notas", "Criado em"]
    ws.append(headers)
    _style_header(ws)

    engine = _conn()
    with engine.connect() as c:
        rows = c.execute(sa.text("""
            SELECT id, entity_type, entity_ref, insurer_name, policy_number, category,
                   coverage_summary, start_date, end_date, renewal_date, payment_frequency,
                   premium_amount, status, auto_renew, notes, created_at
            FROM public.insurance_policies
            ORDER BY status, end_date
        """)).mappings().all()

    today = date.today()
    for r in rows:
        end = r["end_date"]
        days_left = (end - today).days if end else None
        status = r["status"]
        if days_left is not None and days_left < 0:
            status = "expirado"

        ws.append([
            r["id"], r["entity_type"], r["entity_ref"], r["insurer_name"],
            r["policy_number"], r["category"], r["coverage_summary"],
            _fmt(r["start_date"]), _fmt(r["end_date"]), _fmt(r["renewal_date"]),
            r["payment_frequency"], float(r["premium_amount"] or 0),
            status, "Sim" if r["auto_renew"] else "Não",
            r["notes"] or "", _fmt(r["created_at"]),
        ])

    # Date format for cols 8,9,10,16
    for row in ws.iter_rows(min_row=2):
        for i in (7, 8, 9, 15):  # 0-based
            if i < len(row) and isinstance(row[i].value, datetime):
                row[i].number_format = DATE_FORMAT

    _autofit(ws)


def sheet_ideas(wb):
    ws = wb.active
    ws.title = "Ideias"
    headers = ["ID", "Título", "Estado", "Origem", "Criado em"]
    ws.append(headers)
    _style_header(ws)

    engine = _conn()
    with engine.connect() as c:
        rows = c.execute(sa.text("""
            SELECT t.id, t.title, t.status, t.source, t.created_at,
                   COUNT(m.id) AS n_messages
            FROM public.idea_threads t
            LEFT JOIN public.idea_messages m ON m.thread_id = t.id
            GROUP BY t.id
            ORDER BY t.created_at DESC
        """)).mappings().all()

    for r in rows:
        ws.append([r["id"], r["title"], r["status"], r["source"], _fmt(r["created_at"])])

    for row in ws.iter_rows(min_row=2):
        if isinstance(row[4].value, datetime):
            row[4].number_format = "DD/MM/YYYY HH:MM"

    _autofit(ws)


def sheet_decisions(wb):
    ws = wb.active
    ws.title = "Decisões"
    headers = ["ID", "Título", "Tipo", "Estado", "Criado em", "Atualizado em"]
    ws.append(headers)
    _style_header(ws)

    engine = _conn()
    with engine.connect() as c:
        rows = c.execute(sa.text("""
            SELECT id, title, kind, status, created_at, updated_at
            FROM public.decision_queue
            ORDER BY created_at DESC
        """)).mappings().all()

    for r in rows:
        ws.append([r["id"], r["title"], r["kind"], r["status"],
                   _fmt(r["created_at"]), _fmt(r["updated_at"])])

    for row in ws.iter_rows(min_row=2):
        for i in (4, 5):
            if i < len(row) and isinstance(row[i].value, datetime):
                row[i].number_format = "DD/MM/YYYY HH:MM"

    _autofit(ws)


# ── Entry point ───────────────────────────────────────────────────────────────

BUILDERS = {
    "insurance": sheet_insurance,
    "ideas":     sheet_ideas,
    "decisions": sheet_decisions,
}

FILENAMES = {
    "insurance": "seguros",
    "ideas":     "ideias",
    "decisions": "decisoes",
}


def build(export_type: str) -> bytes:
    """Return .xlsx bytes for the given export type."""
    if export_type not in BUILDERS:
        raise ValueError(f"Tipo desconhecido: {export_type}. Disponíveis: {list(BUILDERS)}")
    wb = openpyxl.Workbook()
    BUILDERS[export_type](wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Uso: {sys.argv[0]} <type> [output.xlsx]")
        print(f"Types: {', '.join(BUILDERS)}")
        sys.exit(1)
    export_type = sys.argv[1]
    data = build(export_type)
    if len(sys.argv) >= 3:
        out_path = sys.argv[2]
        with open(out_path, "wb") as f:
            f.write(data)
        print(f"✓ {out_path} ({len(data)//1024}KB)")
    else:
        sys.stdout.buffer.write(data)
